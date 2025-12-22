from __future__ import annotations

from datetime import date, datetime, timedelta
from calendar import monthrange
from io import BytesIO
import csv

from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db import SessionLocal
from app.db_models import CreditoDB, PagamentoDB, AtendenteDB

# Tentamos usar openpyxl. Se não existir, caímos para CSV simples.
try:
    from openpyxl import Workbook  # type: ignore
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore


# ============================================================================
# Helpers comuns
# ============================================================================

def _get_db() -> Session:
    """Cria sessão de BD."""
    return SessionLocal()


def _fmt_kz(valor: float | int | None) -> str:
    v = float(valor or 0.0)
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ") + " Kz"


# ============================================================================
# RELATÓRIOS EM JSON
# ============================================================================

def resumo_geral() -> dict:
    """Totais gerais e indicador de adimplência."""
    db = _get_db()
    try:
        creditos = db.query(CreditoDB).all()

        total_concedido = sum(c.valor_solicitado or 0 for c in creditos)
        total_a_receber = sum(c.valor_total_reembolsar or 0 for c in creditos)
        total_pago = sum(c.valor_pago or 0 for c in creditos)
        total_em_aberto = sum(c.saldo_em_aberto or 0 for c in creditos)

        total_creditos = len(creditos)
        ativos = sum(1 for c in creditos if c.estado == "Ativo")
        devedores = sum(1 for c in creditos if c.estado == "Devedor")
        concluidos = sum(1 for c in creditos if c.estado == "Concluído")

        adimplencia = 0.0
        if total_a_receber > 0:
            adimplencia = (total_pago / total_a_receber) * 100.0

        return {
            "totais": {
                "total_concedido": round(total_concedido, 2),
                "total_a_receber": round(total_a_receber, 2),
                "total_pago": round(total_pago, 2),
                "total_em_aberto": round(total_em_aberto, 2),
                "total_creditos": total_creditos,
                "ativos": ativos,
                "devedores": devedores,
                "concluidos": concluidos,
            },
            "adimplencia": {
                "percentual": round(adimplencia, 2),
            },
            "gerado_em": date.today().isoformat(),
        }
    finally:
        db.close()


def lista_devedores() -> list[dict]:
    """Créditos com saldo em aberto, ordenados do maior saldo."""
    db = _get_db()
    try:
        itens = (
            db.query(CreditoDB)
            .filter(CreditoDB.saldo_em_aberto > 0)
            .order_by(CreditoDB.saldo_em_aberto.desc())
            .all()
        )
        return [
            {
                "id_credito": c.id_credito,
                "nome": c.nome,
                "telefone": c.telefone,
                "valor_solicitado": c.valor_solicitado,
                "valor_total_reembolsar": c.valor_total_reembolsar,
                "valor_pago": c.valor_pago,
                "saldo_em_aberto": c.saldo_em_aberto,
                "data_inicio": c.data_inicio.isoformat(),
                "data_fim": c.data_fim.isoformat(),
                "estado": c.estado,
            }
            for c in itens
        ]
    finally:
        db.close()


def lista_ativos() -> list[dict]:
    db = _get_db()
    try:
        itens = (
            db.query(CreditoDB)
            .filter(CreditoDB.estado == "Ativo")
            .order_by(CreditoDB.id_credito.desc())
            .all()
        )
        return [
            {
                "id_credito": c.id_credito,
                "nome": c.nome,
                "telefone": c.telefone,
                "valor_solicitado": c.valor_solicitado,
                "valor_total_reembolsar": c.valor_total_reembolsar,
                "valor_pago": c.valor_pago,
                "saldo_em_aberto": c.saldo_em_aberto,
                "data_inicio": c.data_inicio.isoformat(),
                "data_fim": c.data_fim.isoformat(),
                "estado": c.estado,
            }
            for c in itens
        ]
    finally:
        db.close()


def lista_concluidos() -> list[dict]:
    db = _get_db()
    try:
        itens = (
            db.query(CreditoDB)
            .filter(CreditoDB.estado == "Concluído")
            .order_by(CreditoDB.id_credito.desc())
            .all()
        )
        return [
            {
                "id_credito": c.id_credito,
                "nome": c.nome,
                "telefone": c.telefone,
                "valor_solicitado": c.valor_solicitado,
                "valor_total_reembolsar": c.valor_total_reembolsar,
                "valor_pago": c.valor_pago,
                "saldo_em_aberto": c.saldo_em_aberto,
                "data_inicio": c.data_inicio.isoformat(),
                "data_fim": c.data_fim.isoformat(),
                "estado": c.estado,
            }
            for c in itens
        ]
    finally:
        db.close()


def top_devedores(limite: int = 10) -> list[dict]:
    db = _get_db()
    try:
        itens = (
            db.query(CreditoDB)
            .filter(CreditoDB.saldo_em_aberto > 0)
            .order_by(CreditoDB.saldo_em_aberto.desc())
            .limit(limite)
            .all()
        )
        return [
            {
                "id_credito": c.id_credito,
                "nome": c.nome,
                "saldo_em_aberto": c.saldo_em_aberto,
            }
            for c in itens
        ]
    finally:
        db.close()


def alertas(dias: int = 7) -> dict:
    """
    - Próximos vencimentos (data_fim entre hoje e hoje+dias, com saldo>0)
    - Devedores (reaproveita lista_devedores)
    """
    hoje = date.today()
    limite = hoje + timedelta(days=dias)

    db = _get_db()
    try:
        proximos = (
            db.query(CreditoDB)
            .filter(
                CreditoDB.saldo_em_aberto > 0,
                CreditoDB.data_fim >= hoje,
                CreditoDB.data_fim <= limite,
            )
            .order_by(CreditoDB.data_fim.asc())
            .all()
        )

        proximos_out = [
            {
                "id_credito": c.id_credito,
                "nome": c.nome,
                "data_fim": c.data_fim.isoformat(),
                "saldo_em_aberto": c.saldo_em_aberto,
                "dias_restantes": (c.data_fim - hoje).days,
            }
            for c in proximos
        ]

        return {
            "proximos_vencimentos": proximos_out,
            "devedores": lista_devedores(),
            "gerado_em": hoje.isoformat(),
        }
    finally:
        db.close()


# ============================================================================
# EXPORTAÇÕES EXCEL / CSV
# ============================================================================

def gerar_resumo_excel() -> StreamingResponse:
    """Gera um Excel simples com o resumo geral."""
    # fallback para CSV se não tiver openpyxl
    if Workbook is None:
        data = resumo_geral()
        buffer = BytesIO()
        writer = csv.writer(buffer, delimiter=";")
        t = data["totais"]
        writer.writerow(["Campo", "Valor"])
        writer.writerow(["Total concedido", t["total_concedido"]])
        writer.writerow(["Total a receber", t["total_a_receber"]])
        writer.writerow(["Total pago", t["total_pago"]])
        writer.writerow(["Total em aberto", t["total_em_aberto"]])
        writer.writerow(["Total créditos", t["total_creditos"]])
        writer.writerow(["Ativos", t["ativos"]])
        writer.writerow(["Devedores", t["devedores"]])
        writer.writerow(["Concluídos", t["concluidos"]])
        buffer.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="resumo_geral.csv"'}
        return StreamingResponse(buffer, media_type="text/csv", headers=headers)

    data = resumo_geral()
    t = data["totais"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    ws.append(["Campo", "Valor"])
    ws.append(["Total concedido", t["total_concedido"]])
    ws.append(["Total a receber", t["total_a_receber"]])
    ws.append(["Total pago", t["total_pago"]])
    ws.append(["Total em aberto", t["total_em_aberto"]])
    ws.append(["Total créditos", t["total_creditos"]])
    ws.append(["Ativos", t["ativos"]])
    ws.append(["Devedores", t["devedores"]])
    ws.append(["Concluídos", t["concluidos"]])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="resumo_geral.xlsx"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def gerar_exportacao_completa_excel() -> StreamingResponse:
    """Excel com abas de créditos e pagamentos."""
    if Workbook is None:
        # fallback: CSV de créditos
        return exportar_creditos_csv()

    db = _get_db()
    try:
        creditos = db.query(CreditoDB).order_by(CreditoDB.id_credito).all()
        pagamentos = db.query(PagamentoDB).order_by(PagamentoDB.id_pagamento).all()
    finally:
        db.close()

    wb = Workbook()
    ws_c = wb.active
    ws_c.title = "Creditos"

    ws_c.append(
        [
            "ID",
            "Nome",
            "Telefone",
            "Profissão",
            "Salário",
            "Valor solicitado",
            "Duração (meses)",
            "Taxa juros",
            "Total a reembolsar",
            "Prestação mensal",
            "Valor pago",
            "Saldo em aberto",
            "Data início",
            "Data fim",
            "Estado",
            "Comentário",
        ]
    )
    for c in creditos:
        ws_c.append(
            [
                c.id_credito,
                c.nome,
                c.telefone,
                c.profissao,
                c.salario_mensal,
                c.valor_solicitado,
                c.duracao_meses,
                c.taxa_juros,
                c.valor_total_reembolsar,
                c.prestacao_mensal,
                c.valor_pago,
                c.saldo_em_aberto,
                c.data_inicio,
                c.data_fim,
                c.estado,
                c.comentario or "",
            ]
        )

    ws_p = wb.create_sheet("Pagamentos")
    ws_p.append(
        [
            "ID pagamento",
            "ID crédito",
            "Nº comprovativo",
            "Data pagamento",
            "Valor pago no dia",
            "Forma pagamento",
            "Observação",
            "ID atendente",
        ]
    )
    for p in pagamentos:
        ws_p.append(
            [
                p.id_pagamento,
                p.id_credito,
                p.nr_comprovativo,
                p.data_pagamento,
                p.valor_pago_no_dia,
                p.forma_pagamento,
                p.observacao or "",
                p.id_atendente,
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {
        "Content-Disposition": 'attachment; filename="ukamba_exportacao_completa.xlsx"'
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def exportar_creditos_csv() -> StreamingResponse:
    """Exporta TODOS os créditos em CSV (separador ';')."""
    db = _get_db()
    try:
        creditos = db.query(CreditoDB).order_by(CreditoDB.id_credito).all()
    finally:
        db.close()

    buffer = BytesIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(
        [
            "id_credito",
            "nome",
            "telefone",
            "profissao",
            "salario_mensal",
            "valor_solicitado",
            "duracao_meses",
            "taxa_juros",
            "valor_total_reembolsar",
            "prestacao_mensal",
            "valor_pago",
            "saldo_em_aberto",
            "data_inicio",
            "data_fim",
            "estado",
            "comentario",
        ]
    )
    for c in creditos:
        writer.writerow(
            [
                c.id_credito,
                c.nome,
                c.telefone,
                c.profissao,
                c.salario_mensal,
                c.valor_solicitado,
                c.duracao_meses,
                c.taxa_juros,
                c.valor_total_reembolsar,
                c.prestacao_mensal,
                c.valor_pago,
                c.saldo_em_aberto,
                c.data_inicio.isoformat(),
                c.data_fim.isoformat(),
                c.estado,
                (c.comentario or "").replace("\n", " "),
            ]
        )
    buffer.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="creditos.csv"'}
    return StreamingResponse(buffer, media_type="text/csv", headers=headers)


def exportar_credito_unico_csv(id_credito: int) -> StreamingResponse:
    """Exporta extrato de um único crédito em CSV."""
    db = _get_db()
    try:
        c = db.query(CreditoDB).filter(CreditoDB.id_credito == id_credito).first()
        if not c:
            buffer = BytesIO()
            writer = csv.writer(buffer, delimiter=";")
            writer.writerow(["erro"])
            writer.writerow([f"Crédito {id_credito} não encontrado"])
            buffer.seek(0)
            headers = {
                "Content-Disposition": f'attachment; filename="credito_{id_credito}_erro.csv"'
            }
            return StreamingResponse(buffer, media_type="text/csv", headers=headers)

        pagamentos = (
            db.query(PagamentoDB)
            .filter(PagamentoDB.id_credito == id_credito)
            .order_by(PagamentoDB.data_pagamento.asc(), PagamentoDB.id_pagamento.asc())
            .all()
        )
    finally:
        db.close()

    buffer = BytesIO()
    writer = csv.writer(buffer, delimiter=";")

    writer.writerow(["Crédito", id_credito])
    writer.writerow(["Nome", c.nome])
    writer.writerow(["Telefone", c.telefone])
    writer.writerow([])
    writer.writerow(
        ["id_pagamento", "data_pagamento", "valor_pago_no_dia", "forma_pagamento", "observacao"]
    )

    for p in pagamentos:
        writer.writerow(
            [
                p.id_pagamento,
                p.data_pagamento.isoformat(),
                p.valor_pago_no_dia,
                p.forma_pagamento,
                (p.observacao or "").replace("\n", " "),
            ]
        )

    buffer.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="credito_{id_credito}_extrato.csv"'
    }
    return StreamingResponse(buffer, media_type="text/csv", headers=headers)


# ============================================================================
# PDFs – relatório mensal & extrato de crédito
# ============================================================================

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm


def _desenhar_cabecalho(c: canvas.Canvas, titulo: str):
    largura, altura = A4
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, altura - 20 * mm, "Ukamba Microcrédito")
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, altura - 26 * mm, titulo)
    c.drawRightString(largura - 20 * mm, altura - 20 * mm, date.today().isoformat())
    c.line(20 * mm, altura - 28 * mm, largura - 20 * mm, altura - 28 * mm)


def relatorio_mensal_pdf(
    ano: int,
    mes: int,
    dias_alerta: int = 7,
    limite_top: int = 10,
    responsavel: str | None = None,
) -> StreamingResponse:
    """
    Gera PDF de resumo mensal.
    Usado em /relatorios/mensal.pdf e pelo botão do dashboard.
    """
    if mes < 1 or mes > 12:
        raise ValueError("Mês inválido (1-12)")
    if ano < 2000 or ano > 2100:
        raise ValueError("Ano inválido")

    inicio_mes = date(ano, mes, 1)
    fim_mes = date(ano, mes, monthrange(ano, mes)[1])

    hoje = date.today()
    limite_alerta = hoje + timedelta(days=dias_alerta)

    db = _get_db()
    try:
        creditos_mes = (
            db.query(CreditoDB)
            .filter(CreditoDB.data_inicio >= inicio_mes, CreditoDB.data_inicio <= fim_mes)
            .order_by(CreditoDB.id_credito)
            .all()
        )

        pagamentos_mes = (
            db.query(PagamentoDB)
            .filter(
                PagamentoDB.data_pagamento >= inicio_mes,
                PagamentoDB.data_pagamento <= fim_mes,
            )
            .order_by(PagamentoDB.data_pagamento, PagamentoDB.id_pagamento)
            .all()
        )

        top_dev = (
            db.query(CreditoDB)
            .filter(CreditoDB.saldo_em_aberto > 0)
            .order_by(CreditoDB.saldo_em_aberto.desc())
            .limit(limite_top)
            .all()
        )

        proximos = (
            db.query(CreditoDB)
            .filter(
                CreditoDB.saldo_em_aberto > 0,
                CreditoDB.data_fim >= hoje,
                CreditoDB.data_fim <= limite_alerta,
            )
            .order_by(CreditoDB.data_fim.asc())
            .all()
        )
    finally:
        db.close()

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    largura, altura = A4

    _desenhar_cabecalho(c, f"Relatório mensal - {mes:02d}/{ano}")

    y = altura - 40 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, "Resumo financeiro do mês")
    y -= 7 * mm

    total_concedido = sum(c_.valor_solicitado or 0 for c_ in creditos_mes)
    total_a_reembolsar = sum(c_.valor_total_reembolsar or 0 for c_ in creditos_mes)
    total_pago_mes = sum(p.valor_pago_no_dia or 0 for p in pagamentos_mes)

    c.setFont("Helvetica", 10)
    c.drawString(22 * mm, y, f"Créditos criados no mês: {len(creditos_mes)}")
    y -= 5 * mm
    c.drawString(22 * mm, y, f"Total concedido no mês: {_fmt_kz(total_concedido)}")
    y -= 5 * mm
    c.drawString(22 * mm, y, f"Total a reembolsar (créditos do mês): {_fmt_kz(total_a_reembolsar)}")
    y -= 5 * mm
    c.drawString(22 * mm, y, f"Pagamentos recebidos no mês: {_fmt_kz(total_pago_mes)}")
    y -= 10 * mm

    # Pagamentos
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, "Pagamentos do mês")
    y -= 6 * mm
    c.setFont("Helvetica", 9)
    c.drawString(22 * mm, y, "Data")
    c.drawString(45 * mm, y, "Crédito")
    c.drawString(70 * mm, y, "Valor")
    c.drawString(95 * mm, y, "Forma")
    y -= 4 * mm
    c.line(20 * mm, y, largura - 20 * mm, y)
    y -= 4 * mm

    for p in pagamentos_mes:
        if y < 30 * mm:
            c.showPage()
            _desenhar_cabecalho(c, f"Relatório mensal - {mes:02d}/{ano}")
            y = altura - 40 * mm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(20 * mm, y, "Pagamentos do mês (cont.)")
            y -= 8 * mm
            c.setFont("Helvetica", 9)

        c.drawString(22 * mm, y, p.data_pagamento.isoformat())
        c.drawString(45 * mm, y, f"#{p.id_credito}")
        c.drawRightString(90 * mm, y, _fmt_kz(p.valor_pago_no_dia))
        c.drawString(95 * mm, y, p.forma_pagamento)
        y -= 4 * mm

    # Nova página: top devedores + próximos vencimentos
    c.showPage()
    _desenhar_cabecalho(c, f"Relatório mensal - {mes:02d}/{ano}")
    y = altura - 40 * mm

    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, "Top devedores")
    y -= 6 * mm
    c.setFont("Helvetica", 9)
    c.drawString(22 * mm, y, "Crédito")
    c.drawString(45 * mm, y, "Nome")
    c.drawRightString(160 * mm, y, "Saldo")
    y -= 4 * mm
    c.line(20 * mm, y, largura - 20 * mm, y)
    y -= 4 * mm

    for d in top_dev:
        if y < 40 * mm:
            c.showPage()
            _desenhar_cabecalho(c, f"Relatório mensal - {mes:02d}/{ano}")
            y = altura - 40 * mm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(20 * mm, y, "Top devedores (cont.)")
            y -= 8 * mm
            c.setFont("Helvetica", 9)

        c.drawString(22 * mm, y, f"#{d.id_credito}")
        c.drawString(45 * mm, y, d.nome)
        c.drawRightString(160 * mm, y, _fmt_kz(d.saldo_em_aberto))
        y -= 4 * mm

    # Próximos vencimentos
    y -= 8 * mm
    if y < 50 * mm:
        c.showPage()
        _desenhar_cabecalho(c, f"Relatório mensal - {mes:02d}/{ano}")
        y = altura - 40 * mm

    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, f"Próximos vencimentos (até {dias_alerta} dias)")
    y -= 6 * mm
    c.setFont("Helvetica", 9)
    c.drawString(22 * mm, y, "Crédito")
    c.drawString(45 * mm, y, "Nome")
    c.drawString(110 * mm, y, "Data fim")
    c.drawRightString(160 * mm, y, "Saldo")
    y -= 4 * mm
    c.line(20 * mm, y, largura - 20 * mm, y)
    y -= 4 * mm

    for cr in proximos:
        if y < 30 * mm:
            c.showPage()
            _desenhar_cabecalho(c, f"Relatório mensal - {mes:02d}/{ano}")
            y = altura - 40 * mm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(20 * mm, y, "Próximos vencimentos (cont.)")
            y -= 8 * mm
            c.setFont("Helvetica", 9)

        c.drawString(22 * mm, y, f"#{cr.id_credito}")
        c.drawString(45 * mm, y, cr.nome)
        c.drawString(110 * mm, y, cr.data_fim.isoformat())
        c.drawRightString(160 * mm, y, _fmt_kz(cr.saldo_em_aberto))
        y -= 4 * mm

    if responsavel:
        c.setFont("Helvetica-Oblique", 9)
        c.drawString(20 * mm, 20 * mm, f"Relatório gerado por: {responsavel}")

    c.save()
    buffer.seek(0)
    filename = f"relatorio_mensal_{ano}_{mes:02d}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


def extrato_credito_pdf(id_credito: int, responsavel: str | None = None) -> StreamingResponse:
    """Extrato em PDF de um único crédito."""
    db = _get_db()
    try:
        c_cred = db.query(CreditoDB).filter(CreditoDB.id_credito == id_credito).first()
        if not c_cred:
            buffer = BytesIO()
            cvs = canvas.Canvas(buffer, pagesize=A4)
            _desenhar_cabecalho(cvs, f"Extrato do crédito #{id_credito}")
            cvs.setFont("Helvetica", 11)
            cvs.drawString(20 * mm, 250 * mm, "Crédito não encontrado.")
            cvs.save()
            buffer.seek(0)
            headers = {
                "Content-Disposition": f'attachment; filename="credito_{id_credito}_nao_encontrado.pdf"'
            }
            return StreamingResponse(buffer, media_type="application/pdf", headers=headers)

        pagamentos = (
            db.query(PagamentoDB)
            .filter(PagamentoDB.id_credito == id_credito)
            .order_by(PagamentoDB.data_pagamento.asc(), PagamentoDB.id_pagamento.asc())
            .all()
        )
    finally:
        db.close()

    buffer = BytesIO()
    cpdf = canvas.Canvas(buffer, pagesize=A4)
    largura, altura = A4

    _desenhar_cabecalho(cpdf, f"Extrato do crédito #{id_credito}")

    y = altura - 40 * mm
    cpdf.setFont("Helvetica-Bold", 11)
    cpdf.drawString(20 * mm, y, f"Cliente: {c_cred.nome}")
    y -= 6 * mm
    cpdf.setFont("Helvetica", 10)
    cpdf.drawString(20 * mm, y, f"Telefone: {c_cred.telefone}")
    y -= 5 * mm
    cpdf.drawString(20 * mm, y, f"Valor solicitado: {_fmt_kz(c_cred.valor_solicitado)}")
    y -= 5 * mm
    cpdf.drawString(20 * mm, y, f"Total a reembolsar: {_fmt_kz(c_cred.valor_total_reembolsar)}")
    y -= 5 * mm
    cpdf.drawString(20 * mm, y, f"Valor pago: {_fmt_kz(c_cred.valor_pago)}")
    y -= 5 * mm
    cpdf.drawString(20 * mm, y, f"Saldo em aberto: {_fmt_kz(c_cred.saldo_em_aberto)}")
    y -= 10 * mm

    cpdf.setFont("Helvetica-Bold", 11)
    cpdf.drawString(20 * mm, y, "Pagamentos")
    y -= 6 * mm
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(22 * mm, y, "Data")
    cpdf.drawString(45 * mm, y, "Comprovativo")
    cpdf.drawRightString(110 * mm, y, "Valor")
    cpdf.drawString(120 * mm, y, "Forma")
    y -= 4 * mm
    cpdf.line(20 * mm, y, largura - 20 * mm, y)
    y -= 4 * mm

    for p in pagamentos:
        if y < 30 * mm:
            cpdf.showPage()
            _desenhar_cabecalho(cpdf, f"Extrato do crédito #{id_credito}")
            y = altura - 40 * mm
            cpdf.setFont("Helvetica-Bold", 11)
            cpdf.drawString(20 * mm, y, "Pagamentos (cont.)")
            y -= 8 * mm
            cpdf.setFont("Helvetica", 9)

        cpdf.drawString(22 * mm, y, p.data_pagamento.isoformat())
        cpdf.drawString(45 * mm, y, p.nr_comprovativo)
        cpdf.drawRightString(110 * mm, y, _fmt_kz(p.valor_pago_no_dia))
        cpdf.drawString(120 * mm, y, p.forma_pagamento)
        y -= 4 * mm

    if responsavel:
        cpdf.setFont("Helvetica-Oblique", 9)
        cpdf.drawString(20 * mm, 20 * mm, f"Extrato emitido por: {responsavel}")

    cpdf.save()
    buffer.seek(0)
    filename = f"credito_{id_credito}_extrato.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)
